# -*- coding: utf-8 -*-
import json, datetime, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = "https://nara-career.com"
OUT_DIR = r"C:\Users\naoki\AIエージェント用\副業検討"

# ---- load data ----
ART = json.load(open('pages_meta.json', encoding='utf-8'))   # articles + index pages (has 'type')
WK = json.load(open('ga4_weekly.json', encoding='utf-8'))
CUM = json.load(open('ga4_cumulative.json', encoding='utf-8'))
try:
    GSC = json.load(open('gsc_weekly.json', encoding='utf-8'))
except FileNotFoundError:
    GSC = []

# ---- period params (params.json overrides; else compute from today) ----
def _d(s):
    try:
        return datetime.date.fromisoformat(s)
    except Exception:
        return None

today = datetime.date.today()
this_mon = today - datetime.timedelta(days=today.weekday())
defaults = {
    "week_start": str(this_mon - datetime.timedelta(days=7)),
    "week_end":   str(this_mon - datetime.timedelta(days=1)),
    "cum_start":  "2026-01-01",
    "cum_end":    str(today - datetime.timedelta(days=1)),
    "run_date":   str(today),
}
P = dict(defaults)
if os.path.exists('params.json'):
    P.update(json.load(open('params.json', encoding='utf-8')))
ws_d, we_d = _d(P["week_start"]), _d(P["week_end"])
run_d = _d(P["run_date"])
md = lambda x: f"{x.month}/{x.day}"

def lastseg(url):
    p = url.replace(BASE, "").rstrip("/")
    return "_home" if p == "" else p.split("/")[-1]

# PV maps keyed by last path segment (handles old/new URL variants of same page)
def pvmap(d):
    m = {}
    for path, v in d.items():
        k = "_home" if path.rstrip("/") == "" else path.rstrip("/").split("/")[-1]
        m[k] = m.get(k, 0) + v
    return m

WKm, CUMm = pvmap(WK), pvmap(CUM)

# GSC map keyed by slug (aggregates old/new + www/non-www; position = impression-weighted avg)
def gscmap(rows):
    agg = {}
    for r in rows:
        p = r["url"].split("//", 1)[-1]
        p = "/" + p.split("/", 1)[-1] if "/" in p else "/"
        k = "_home" if p.rstrip("/") == "" else p.rstrip("/").split("/")[-1]
        a = agg.setdefault(k, {"clicks": 0, "impressions": 0, "pos_w": 0.0})
        a["clicks"] += r.get("clicks", 0)
        a["impressions"] += r.get("impressions", 0)
        a["pos_w"] += r.get("position", 0) * r.get("impressions", 0)
    out = {}
    for k, a in agg.items():
        imp = a["impressions"]
        out[k] = {"clicks": a["clicks"], "impressions": imp,
                  "ctr": (a["clicks"] / imp) if imp else 0,
                  "position": (a["pos_w"] / imp) if imp else None}
    return out

GSCm = gscmap(GSC)

# ---- assemble rows ----
rows = []
for a in ART:
    k = lastseg(a["url"])
    rows.append({
        "type": a.get("type", "記事"), "cat": a["category"], "title": a["title"],
        "slug": a["slug"], "url": a["url"],
        "date": _d(a.get("date", "")), "updated": _d(a["updated"]) if a.get("updated") else None,
        "uc": a.get("updateCount", 0), "pv_cum": CUMm.get(k, 0), "pv_wk": WKm.get(k, 0),
    })

CAT_ORDER = ["共通","career","tenshoku","mindset","story","sidejob","taishoku","shitsugyo","shikaku","coaching","english","blog","sim","hikari","kaikei","furusato"]
order = {c:i for i,c in enumerate(CAT_ORDER)}
rows.sort(key=lambda r: (order.get(r["cat"], 99), r["date"] or datetime.date(2100,1,1), r["type"]!="インデックス"))

# ---- write xlsx ----
wb = Workbook(); ws = wb.active; ws.title = "ページ一覧"
headers = ["No.","種別","カテゴリ","ページタイトル","slug","URL","公開日","更新日","更新回数",
           "クリック数\n(週次GSC)","表示回数\n(週次GSC)","CTR\n(週次GSC)","掲載順位\n(週次GSC)",
           "PV累計", f"週次PV\n({md(ws_d)}-{md(we_d)})"]

ws.merge_cells("A1:O1")
ws["A1"] = f"nara-career.com ページ分析  (作成: {P['run_date']})"
ws["A1"].font = Font(bold=True, size=14)
ws.merge_cells("A2:O2")
ws["A2"] = (f"週次GSC/PV={P['week_start']}〜{P['week_end']}(月〜日) / PV累計={P['cum_start']}〜{P['cum_end']} / "
            "PV・GSCはともにURL再構築の旧/新(www含む)URLをslugで名寄せ集計(GSC順位は表示回数加重平均) / "
            "更新回数=gitコミット数-1 / 空欄=その週GSC表示0")
ws["A2"].font = Font(size=9, color="C00000")
ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
ws.row_dimensions[2].height = 30

HROW = 4
hfill = PatternFill("solid", fgColor="1F4E78")
hfont = Font(bold=True, color="FFFFFF", size=10)
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for j, h in enumerate(headers, 1):
    c = ws.cell(HROW, j, h)
    c.fill = hfill; c.font = hfont; c.border = border
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[HROW].height = 34

artfill = PatternFill("solid", fgColor="FFFFFF")
idxfill = PatternFill("solid", fgColor="FCE4D6")
for i, r in enumerate(rows, 1):
    rr = HROW + i
    g = GSCm.get(lastseg(r["url"]))
    if g:
        g_clicks, g_imp, g_ctr = g["clicks"], g["impressions"], g["ctr"]
        g_pos = round(g["position"], 1) if g["position"] is not None else None
    else:
        g_clicks = g_imp = g_ctr = g_pos = None
    vals = [i, r["type"], r["cat"], r["title"], r["slug"], r["url"], r["date"], r["updated"], r["uc"],
            g_clicks, g_imp, g_ctr, g_pos, r["pv_cum"], r["pv_wk"]]
    for j, v in enumerate(vals, 1):
        c = ws.cell(rr, j, v)
        c.border = border
        c.fill = idxfill if r["type"] == "インデックス" else artfill
        if j in (7,8):
            c.number_format = "yyyy-mm-dd"
        if j == 12 and v is not None:
            c.number_format = "0.0%"
        if j == 6:
            c.font = Font(color="0563C1", size=9)
            if r["url"]:
                c.hyperlink = r["url"]
        if j in (1,9,10,11,12,13,14,15):
            c.alignment = Alignment(horizontal="center")

widths = [5,11,10,46,26,42,12,12,8,11,11,9,10,9,11]
for j, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(j)].width = w

ws.freeze_panes = "A5"
ws.auto_filter.ref = f"A{HROW}:O{HROW+len(rows)}"

stamp = run_d.strftime("%Y%m%d") if run_d else P["run_date"].replace("-", "")
out = os.path.join(OUT_DIR, f"nara-career_ページ分析_{stamp}.xlsx")
wb.save(out)
print("rows:", len(rows), "articles:", sum(1 for r in rows if r['type']=='記事'),
      "index:", sum(1 for r in rows if r['type']=='インデックス'))
print("saved:", out)
